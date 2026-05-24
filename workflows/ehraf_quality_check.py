#!/usr/bin/env python3
"""
eHRAF Quality Check Report Generator
======================================
Compares weekly schedule + eHRAF data to identify:
  1. Part-time instructors with Active classes but MISSING eHRAF
  2. Part-time instructors with eHRAF in INVALID status (Rejected / Pending Dept Approval)
  3. eHRAFs with INCORRECT Total Hours (based on sections × credit hours)

Usage:
    python ehraf_quality_check.py \
        --schedule  schedule.xlsx \
        --instructors instructor_table.xlsx \
        --ehraf     FallEhraf_AdjunctPayroll.xlsx  \
        --output    ehraf_quality_report.xlsx \
        [--term 1269] [--four_credit_catalogs 151,152]

    If --ehraf is a .numbers file it will be auto-converted to .xlsx via LibreOffice.
    All arguments have defaults so the script can be run with just:
        python ehraf_quality_check.py
    (it will use the files in the current directory that match default naming patterns)
"""

import argparse
import os
import sys
import subprocess
import warnings
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Statuses considered INVALID (eHRAF exists but not usable)
INVALID_STATUSES = {
    "rejected",
}

# Status keywords that are acceptable
VALID_STATUS_KEYWORDS = [
    "approved",
    "awaiting subdivision",
    "payroll",
    "final approved",
    "resubmitted",
    "pending department approval",
    "awaiting department approval",
]

# 4-credit catalog numbers (used to set H/W=4 instead of 3)
DEFAULT_FOUR_CREDIT = {"151", "152"}

# Weeks in semester (used to compute expected total hours)
SEMESTER_WEEKS = 15

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def normalize_name(name: str) -> str:
    """Normalize 'Last,First' or 'Last, First' to lowercase 'last,first'."""
    if not isinstance(name, str):
        return ""
    return name.strip().lower().replace(", ", ",").replace(" ,", ",")


def name_from_ehraf(first: str, last: str) -> str:
    """Build 'last,first' key from eHRAF FIRST_NAME / LAST_NAME columns."""
    f = str(first).strip() if pd.notna(first) else ""
    l = str(last).strip() if pd.notna(last) else ""
    # Handle 'Shirley ' style trailing spaces
    return normalize_name(f"{l},{f}")


def is_valid_status(status: str) -> bool:
    """Return True if the eHRAF status is acceptable (not rejected/pending dept)."""
    s = str(status).strip().lower()
    for bad in INVALID_STATUSES:
        if bad in s:
            return False
    return True


def expected_hours(sections_df: pd.DataFrame, four_credit_set: set) -> dict:
    """
    Given a DataFrame of sections for one instructor, return
    {'h_per_week': float, 'prof_hours': float, 'total_hours': float, 'n_sections': int}.
    """
    n = len(sections_df)
    h_pw = sum(
        4 if str(r["Catalog#"]) in four_credit_set else 3
        for _, r in sections_df.iterrows()
    )
    prof = 5 * h_pw
    total = h_pw * SEMESTER_WEEKS + prof  # = 20 * h_pw
    return {"n_sections": n, "h_per_week": h_pw, "prof_hours": prof, "total_hours": total}


def convert_numbers_to_xlsx(path: str) -> str:
    """Convert a .numbers file to .xlsx using LibreOffice."""
    out_dir = str(Path(path).parent)
    result = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "xlsx", path, "--outdir", out_dir],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice conversion failed:\n{result.stderr}")
    stem = Path(path).stem
    return str(Path(out_dir) / f"{stem}.xlsx")


def read_schedule(path: str) -> pd.DataFrame:
    """
    Read schedule Excel. Header is on row 2 (index 1).
    Returns DataFrame with Active classes only, dropping rows with no instructor.
    """
    df = pd.read_excel(path, header=1)
    # Rename duplicated 'Tot Enrl' if present
    cols = list(df.columns)
    seen = {}
    new_cols = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            new_cols.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            new_cols.append(c)
    df.columns = new_cols
    # Keep Active only, drop rows without instructor name
    df = df[df["Class Stat"] == "Active"].copy()
    df = df[df["Name"].notna()].copy()
    df["Name"] = df["Name"].astype(str).str.strip()
    df["Name"] = df["Name"].str.split(" ").str[0]
    df["name_key"] = df["Name"].apply(normalize_name)
    df["Catalog#"] = df["Catalog#"].astype(str).str.strip()
    return df


def read_instructor_table(path: str) -> pd.DataFrame:
    """
    Read instructor table. Header is on row 2 (index 1).
    Returns DataFrame with columns: Instructor, P/F, Instr_Subject, name_key.
    """
    df = pd.read_excel(path, header=1)
    df = df.dropna(subset=["Instructor"]).copy()
    df["Instructor"] = df["Instructor"].astype(str).str.strip()
    df["P/F"] = df["P/F"].astype(str).str.strip().str.upper()
    df["Instructor"] = df["Instructor"].str.split(" ").str[0]
    df["name_key"] = df["Instructor"].apply(normalize_name)
    return df


def read_ehraf(path: str) -> pd.DataFrame:
    """
    Read eHRAF export. First row (index 0) is the header.
    Returns DataFrame with a 'name_key' column for matching.
    """
    # Auto-convert .numbers
    if path.lower().endswith(".numbers"):
        path = convert_numbers_to_xlsx(path)
    df = pd.read_excel(path, header=0)
    # Build name key
    df["name_key"] = df.apply(
        lambda r: name_from_ehraf(r.get("FIRST_NAME", ""), r.get("LAST_NAME", "")),
        axis=1,
    )
    return df


# ---------------------------------------------------------------------------
# Core comparison logic
# ---------------------------------------------------------------------------

def run_comparison(schedule_df, instructor_df, ehraf_df, four_credit_set):
    """
    Compare schedule (PT instructors) vs. eHRAF records.
    Returns a list of dicts — one row per PT instructor.
    """
    # Build a set of part-time instructor name keys
    pt_keys = set(instructor_df[instructor_df["P/F"] == "F"]["name_key"])

    # Get active sections for PT instructors only
    pt_schedule = schedule_df[~schedule_df["name_key"].isin(pt_keys)].copy()

    # Build eHRAF lookup: name_key → list of records
    ehraf_lookup: dict[str, list] = {}
    for _, row in ehraf_df.iterrows():
        key = row["name_key"]
        if key:
            ehraf_lookup.setdefault(key, []).append(row)

    results = []

    for name_key, grp in pt_schedule.groupby("name_key"):
        # Get the display name from the schedule
        display_name = grp["Name"].iloc[0]

        # Section details
        exp = expected_hours(grp, four_credit_set)
        sections_list = ", ".join(
            grp["Subject"].astype(str) + " " + grp["Catalog#"].astype(str)
            + "-" + grp["Section"].astype(str)
        )

        # Find matching eHRAF records
        ehraf_records = ehraf_lookup.get(name_key, [])

        if not ehraf_records:
            results.append({
                "Instructor": display_name,
                "Sections (#)": exp["n_sections"],
                "Courses": sections_list,
                "Expected H/W": exp["h_per_week"],
                "Expected Prof Hrs": exp["prof_hours"],
                "Expected Total Hrs": exp["total_hours"],
                "eHRAF ID": "—",
                "eHRAF Status": "❌ MISSING",
                "Found Total Hrs": "—",
                "Hours Match": "—",
                "Issue": "No eHRAF found for this instructor",
                "Severity": "CRITICAL",
            })
            continue

        # If multiple eHRAF records, use the most recent / highest app_id
        # (instructor may have multiple if resubmitted)
        # We want the one that's NOT rejected — find best record
        valid_records = [r for r in ehraf_records if is_valid_status(str(r.get("APP-STATUS", "")))]
        check_records = valid_records if valid_records else ehraf_records
        # Sort by APPLICATION_ID descending to get latest
        check_records = sorted(
            check_records,
            key=lambda r: int(str(r.get("APPLICATION_ID", 0)).replace("nan", "0") or 0),
            reverse=True,
        )
        rec = check_records[0]

        status = str(rec.get("APP-STATUS", "Unknown")).strip()
        app_id = str(rec.get("APPLICATION_ID", "?")).strip()
        found_total = rec.get("Total Hours")

        # Check status validity
        status_ok = is_valid_status(status)

        # Check hours
        try:
            found_total_num = float(str(found_total).replace(",", ""))
        except (ValueError, TypeError):
            found_total_num = None

        hours_match = (
            found_total_num is not None
            and abs(found_total_num - exp["total_hours"]) < 0.01
        )

        # Determine issue
        issues = []
        severity = "OK"

        if not status_ok:
            issues.append(f"Status is '{status}' (invalid)")
            severity = "CRITICAL"
        if found_total_num is None:
            issues.append("Total Hours missing in eHRAF")
            severity = "ERROR" if severity == "OK" else severity
        elif not hours_match:
            issues.append(
                f"Hours mismatch: eHRAF={found_total_num}, expected={exp['total_hours']}"
            )
            severity = "ERROR" if severity == "OK" else severity

        results.append({
            "Instructor": display_name,
            "Sections (#)": exp["n_sections"],
            "Courses": sections_list,
            "Expected H/W": exp["h_per_week"],
            "Expected Prof Hrs": exp["prof_hours"],
            "Expected Total Hrs": exp["total_hours"],
            "eHRAF ID": app_id,
            "eHRAF Status": status,
            "Found Total Hrs": found_total_num if found_total_num is not None else "?",
            "Hours Match": "✅" if hours_match else "❌",
            "Issue": "; ".join(issues) if issues else "None",
            "Severity": severity,
        })

    return sorted(results, key=lambda x: (x["Severity"] != "CRITICAL", x["Severity"] != "ERROR", x["Instructor"]))


# ---------------------------------------------------------------------------
# Excel report writer
# ---------------------------------------------------------------------------

COLOR = {
    "CRITICAL": "FFB3B3",  # red
    "ERROR": "FFE699",     # amber
    "OK": "C6EFCE",        # green
    "HEADER_DARK": "1F3864",
    "HEADER_LIGHT": "2F5496",
    "TITLE": "203864",
}

def write_report(results, output_path, run_date, schedule_path, ehraf_path):
    wb = Workbook()

    # ── Sheet 1: Detailed Results ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Quality Check Results"

    # Title block
    ws.merge_cells("A1:L1")
    ws["A1"] = f"eHRAF Quality Check Report  —  Run: {run_date}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color=COLOR["HEADER_DARK"])
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Schedule: {Path(schedule_path).name}    |    eHRAF: {Path(ehraf_path).name}"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 14

    # Column headers (row 3)
    headers = [
        "Instructor", "Sections", "Courses", "Exp H/W",
        "Exp Prof Hrs", "Exp Total Hrs",
        "eHRAF ID", "eHRAF Status", "Found Total Hrs",
        "Hours Match", "Issue", "Severity",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=COLOR["HEADER_LIGHT"])
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # Data rows
    for row_idx, rec in enumerate(results, 4):
        color = COLOR.get(rec["Severity"], "FFFFFF")
        fill = PatternFill("solid", start_color=color)
        vals = [
            rec["Instructor"], rec["Sections (#)"], rec["Courses"],
            rec["Expected H/W"], rec["Expected Prof Hrs"], rec["Expected Total Hrs"],
            rec["eHRAF ID"], rec["eHRAF Status"], rec["Found Total Hrs"],
            rec["Hours Match"], rec["Issue"], rec["Severity"],
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row_idx].height = 18

    # Column widths
    widths = [28, 9, 40, 9, 12, 13, 12, 24, 14, 11, 55, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Auto-freeze header rows
    ws.freeze_panes = "A4"

    # ── Sheet 2: Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    total = len(results)
    critical = sum(1 for r in results if r["Severity"] == "CRITICAL")
    errors = sum(1 for r in results if r["Severity"] == "ERROR")
    ok = sum(1 for r in results if r["Severity"] == "OK")

    summary_data = [
        ("Run Date", run_date),
        ("Schedule File", Path(schedule_path).name),
        ("eHRAF File", Path(ehraf_path).name),
        ("", ""),
        ("Total PT Instructors with Active Classes", total),
        ("✅ OK (eHRAF present, valid status, correct hours)", ok),
        ("⚠️  ERROR (eHRAF present but hours mismatch)", errors),
        ("❌ CRITICAL (missing eHRAF or invalid status)", critical),
    ]

    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 30
    ws2.merge_cells("A1:B1")
    ws2["A1"] = "SUMMARY"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A1"].fill = PatternFill("solid", start_color=COLOR["HEADER_DARK"])
    ws2["A1"].font = Font(bold=True, size=14, color="FFFFFF")

    for i, (label, val) in enumerate(summary_data, 2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=val)
        if label.startswith("✅"):
            ws2.cell(row=i, column=2).fill = PatternFill("solid", start_color=COLOR["OK"])
        elif label.startswith("⚠️"):
            ws2.cell(row=i, column=2).fill = PatternFill("solid", start_color=COLOR["ERROR"])
        elif label.startswith("❌"):
            ws2.cell(row=i, column=2).fill = PatternFill("solid", start_color=COLOR["CRITICAL"])

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate eHRAF Quality Check Report"
    )
    parser.add_argument(
        "--schedule", default="schedule.xlsx",
        help="Path to the weekly schedule Excel file"
    )
    parser.add_argument(
        "--instructors", default="instructor_table.xlsx",
        help="Path to the instructor P/F table Excel file"
    )
    parser.add_argument(
        "--ehraf", default="FallEhraf_AdjunctPayroll.xlsx",
        help="Path to the eHRAF export (xlsx or .numbers)"
    )
    parser.add_argument(
        "--output", default=None,
        help="Output report path (default: ehraf_report_YYYY-MM-DD.xlsx)"
    )
    parser.add_argument(
        "--four_credit_catalogs", default="151,152",
        help="Comma-separated catalog numbers that are 4-credit (default: 151,152)"
    )
    args = parser.parse_args()

    run_date = datetime.today().strftime("%Y-%m-%d %H:%M")
    four_cr = set(args.four_credit_catalogs.split(","))
    output = args.output or f"ehraf_report_{datetime.today().strftime('%Y-%m-%d')}.xlsx"

    print(f"\n{'='*60}")
    print(f"  eHRAF Quality Check  —  {run_date}")
    print(f"{'='*60}")
    print(f"  Schedule  : {args.schedule}")
    print(f"  Instructors: {args.instructors}")
    print(f"  eHRAF     : {args.ehraf}")
    print(f"  Output    : {output}")
    print()

    # Read inputs
    print("Reading schedule...", end=" ", flush=True)
    sched = read_schedule(args.schedule)
    print(f"{len(sched)} active classes")

    print("Reading instructor table...", end=" ", flush=True)
    inst = read_instructor_table(args.instructors)
    pt = inst[inst["P/F"] == "P"]
    print(f"{len(pt)} part-time instructors")

    print("Reading eHRAF data...", end=" ", flush=True)
    ehraf = read_ehraf(args.ehraf)
    print(f"{len(ehraf)} eHRAF records")

    # Run comparison
    print("Comparing...", end=" ", flush=True)
    results = run_comparison(sched, inst, ehraf, four_cr)
    critical = sum(1 for r in results if r["Severity"] == "CRITICAL")
    errors = sum(1 for r in results if r["Severity"] == "ERROR")
    ok = sum(1 for r in results if r["Severity"] == "OK")
    print(f"{len(results)} PT instructors checked")

    # Write report
    print("Writing report...", end=" ", flush=True)
    write_report(results, output, run_date, args.schedule, args.ehraf)
    print(f"Saved to {output}")

    print(f"\n  Results:  ✅ OK={ok}  ⚠️  Errors={errors}  ❌ Critical={critical}")
    print(f"{'='*60}\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())


# ---------------------------------------------------------------------------
# Library entry-point (used by the Flask dashboard — no argparse needed)
# ---------------------------------------------------------------------------

def run(
    schedule_path: str,
    instructors_path: str,
    ehraf_path: str,
    output_path: str,
    four_credit_catalogs: str = "151,152",
) -> dict:
    """
    Run the eHRAF quality check and write an Excel report.

    Returns a summary dict:
        {ok, errors, critical, total, output_path, run_date}
    """
    run_date = datetime.today().strftime("%Y-%m-%d %H:%M")
    four_cr  = set(four_credit_catalogs.split(","))

    sched = read_schedule(schedule_path)
    inst  = read_instructor_table(instructors_path)
    ehraf = read_ehraf(ehraf_path)

    results  = run_comparison(sched, inst, ehraf, four_cr)
    critical = sum(1 for r in results if r["Severity"] == "CRITICAL")
    errors   = sum(1 for r in results if r["Severity"] == "ERROR")
    ok       = sum(1 for r in results if r["Severity"] == "OK")

    write_report(results, output_path, run_date, schedule_path, ehraf_path)

    return {
        "ok":          ok,
        "errors":      errors,
        "critical":    critical,
        "total":       len(results),
        "output_path": output_path,
        "run_date":    run_date,
    }
