"""
ehraf_generator.py — Generate eHRAF adjunct payroll spreadsheet.

Summer/Winter: all instructors in schedule.
Fall/Spring:   only part-time instructors (P/F == 'P' in instructor table).

Hours formula (standard CUNY):
  Teaching Hours = Credits × 15
  PD Hours       = 15  (flat)
  Total          = Teaching + PD
"""

import io

import openpyxl
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter
import pandas as pd

# ── Course credit overrides (default is 3) ───────────────────────────────────

# Calculus sequence only — pre-calculus (104/105/106) and MAT 301 are 3 credits.
FOUR_CREDIT_COURSES = {
    ('MAT', '151'),   # Calculus I
    ('MAT', '152'),   # Calculus II
    ('MAT', '241'),   # Calculus III / Multivariable
}


def _credits(subject: str, catalog: str) -> int:
    return 4 if (subject.strip().upper(), catalog.strip()) in FOUR_CREDIT_COURSES else 3


def _norm_name(name: str) -> tuple:
    """Return (last_lower, first_lower) for cross-semester comparison."""
    name = name.strip()
    if ',' in name:
        parts = name.split(',', 1)
        last  = parts[0].strip().lower()
        first = parts[1].strip().lower().split()[0] if parts[1].strip() else ''
    else:
        parts = name.split()
        last  = parts[-1].lower() if parts else ''
        first = parts[0].lower()  if len(parts) >= 2 else ''
    return (last, first)


def _appointment_type(instructor: str, pf: str, season: str, prev_names: set) -> str:
    """Return 'New Hire' or 'Repeat Appointment'.

    Full-time instructors teaching Winter/Summer are always New Hire (extra teaching
    outside their regular contract). All others are Repeat if they appear in the
    previous semester's schedule, otherwise New Hire.
    """
    if pf == 'F' and season in ('Winter', 'Summer'):
        return 'New Hire'
    return 'Repeat Appointment' if _norm_name(instructor) in prev_names else 'New Hire'


# ── File readers ──────────────────────────────────────────────────────────────

def _read_schedule(data: bytes) -> pd.DataFrame:
    """
    Read a schedule export (summer or fall format).
    Row 0 is a junk title row; row 1 is the real header.
    """
    df = pd.read_excel(io.BytesIO(data), header=1, dtype=str)
    df.columns = df.columns.str.strip()

    # Keep only Active sections with an assigned instructor
    df = df[df['Class Stat'].str.strip() == 'Active']
    df = df[df['Name'].notna() & (df['Name'].str.strip() != '')]

    for col in ('Subject', 'Catalog#', 'Section', 'Class Title', 'Name', 'Session'):
        if col in df.columns:
            df[col] = df[col].fillna('').astype(str).str.strip()

    # Zero-pad section numbers that are purely numeric
    df['Section'] = df['Section'].apply(
        lambda s: s.zfill(2) if s.isdigit() else s
    )
    return df.reset_index(drop=True)


def _read_instructor_table(data: bytes) -> pd.DataFrame:
    """
    Read instructor_table.xlsx.
    Row 0 blank, row 1 = header: Instructor | P/F | Instr_Subject
    """
    df = pd.read_excel(io.BytesIO(data), header=1, dtype=str)
    df.columns = df.columns.str.strip()
    df = df[df['Instructor'].notna() & (df['Instructor'].str.strip() != '')]
    df['Instructor'] = df['Instructor'].str.strip()
    df['P/F'] = df['P/F'].fillna('').str.strip().str.upper()
    return df.reset_index(drop=True)


# ── Excel styling helpers ─────────────────────────────────────────────────────

_NAVY    = '1F3864'
_BLUE    = 'D9E1F2'
_YELLOW  = 'FFF2CC'
_WHITE   = 'FFFFFF'
_LGRAY   = 'F5F5F5'

_thin   = Side(style='thin',   color='BBBBBB')
_medium = Side(style='medium', color='888888')
_THIN_B   = Border(left=_thin,   right=_thin,   top=_thin,   bottom=_thin)
_HEADER_B = Border(left=_medium, right=_medium, top=_medium, bottom=_medium)


def _fill(color: str) -> PatternFill:
    return PatternFill('solid', fgColor=color)


def _set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    return c


# ── Main generator ────────────────────────────────────────────────────────────

def generate(schedule_bytes: bytes,
             instructor_bytes: bytes,
             season: str,
             year: int,
             sessions: list[dict],
             prev_schedule_bytes: bytes | None = None) -> tuple[bytes, list]:
    """
    Parameters
    ----------
    season          : 'Summer' | 'Winter' | 'Fall' | 'Spring'
    year            : e.g. 2026
    sessions        : list of {'code': str, 'start': str, 'end': str}
                      Summer/Winter: one entry per sub-session (5W1, 5W2 …)
                      Fall/Spring: single entry, code ignored

    Returns
    -------
    bytes — xlsx file ready for download
    """
    sched = _read_schedule(schedule_bytes)
    instr = _read_instructor_table(instructor_bytes)

    # Build previous-semester instructor set for appointment type determination
    prev_names: set = set()
    if prev_schedule_bytes:
        try:
            prev_df = _read_schedule(prev_schedule_bytes)
            prev_names = {_norm_name(n) for n in prev_df['Name'].dropna() if str(n).strip()}
        except Exception:
            pass

    # P/F lookup for full-time detection
    pf_lookup = dict(zip(instr['Instructor'], instr['P/F']))

    is_full_semester = season in ('Fall', 'Spring')

    # ── Filter by session codes (always) ─────────────────────────────────────
    session_codes = {s['code'] for s in sessions if s.get('code')}
    if session_codes and 'Session' in sched.columns:
        sched = sched[sched['Session'].isin(session_codes)]

    # ── Filter instructors ────────────────────────────────────────────────────
    if is_full_semester:
        pt_names = set(instr[instr['P/F'] == 'P']['Instructor'])
        sched = sched[sched['Name'].isin(pt_names)]

    # ── Build session date lookup ─────────────────────────────────────────────
    sess_map = {s['code']: s for s in sessions}

    # ── Sort: instructor → subject → catalog → section ───────────────────────
    sort_cols = [c for c in ('Name', 'Subject', 'Catalog#', 'Section')
                 if c in sched.columns]
    sched = sched.sort_values(sort_cols).reset_index(drop=True)

    # ── Build row data ────────────────────────────────────────────────────────
    records = []
    counter = 1

    for instructor, grp in sched.groupby('Name', sort=True):
        sub_credits = sub_teach = sub_pd = sub_total = 0
        section_count = 0

        pf = pf_lookup.get(instructor, 'P')
        appt = _appointment_type(instructor, pf, season, prev_names)

        for _, row in grp.iterrows():
            subj    = row.get('Subject', '')
            cat     = row.get('Catalog#', '')
            sect    = row.get('Section', '')
            title   = row.get('Class Title', '')
            sess_cd = row.get('Session', '') if not is_full_semester else ''

            cr  = _credits(subj, cat)
            th  = cr * 15
            pd_ = 15
            tot = th + pd_

            records.append({
                'type':             'course',
                'num':              counter,
                'instructor':       instructor,
                'appointment_type': appt,
                'session':          sess_cd,
                'subject':          subj,
                'catalog':          cat,
                'section':          sect,
                'course_code':      f'{subj} {cat}-{sect}',
                'title':            title,
                'credits':          cr,
                'teaching':         th,
                'pd':               pd_,
                'total':            tot,
            })
            sub_credits += cr; sub_teach += th; sub_pd += pd_; sub_total += tot
            section_count += 1
            counter += 1

        records.append({
            'type':             'subtotal',
            'num':              '',
            'instructor':       f'Subtotal — {instructor}',
            'appointment_type': appt,
            'session':          '',
            'subject':          '',
            'catalog':          '',
            'section':          '',
            'course_code':      '',
            'title':            f'Sections: {section_count}',
            'credits':          sub_credits,
            'teaching':         sub_teach,
            'pd':               sub_pd,
            'total':            sub_total,
        })

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{season} {year}'

    # Show session column if there are multiple distinct sessions
    include_session_col = len(sessions) > 1

    if include_session_col:
        COLS = ['#', 'Instructor Name', 'Appointment Type', 'Session',
                'Subject', 'Catalog #', 'Section', 'Course Code', 'Class Title',
                'Credits', 'Teaching Hours', 'PD Hours', 'Total Hours']
    else:
        COLS = ['#', 'Instructor Name', 'Appointment Type',
                'Subject', 'Catalog #', 'Section', 'Course Code', 'Class Title',
                'Credits', 'Teaching Hours', 'PD Hours', 'Total Hours']

    N  = len(COLS)
    LC = get_column_letter(N)

    # Row 1 — title
    ws.merge_cells(f'A1:{LC}1')
    ws['A1'] = f'{season} {year} EHRAF Payroll List'
    ws['A1'].font      = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill      = _fill(_NAVY)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Row 2 — date info
    ws.merge_cells(f'A2:{LC}2')
    if is_full_semester and sessions:
        s = sessions[0]
        date_text = f"Semester Dates: {s['start']}  →  {s['end']}"
    elif sessions:
        parts = [f"{s['code']}: {s['start']} – {s['end']}"
                 for s in sessions if s.get('code')]
        date_text = 'Sessions:   ' + '     |     '.join(parts)
    else:
        date_text = ''
    ws['A2'] = date_text
    ws['A2'].font      = Font(italic=True, size=10, color='444444')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16

    # Row 3 — blank spacer
    ws.row_dimensions[3].height = 6

    # Row 4 — column headers
    hdr_font  = Font(bold=True, size=10, color='FFFFFF')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for ci, name in enumerate(COLS, 1):
        _set_cell(ws, 4, ci, name,
                  font=hdr_font, fill=_fill(_NAVY),
                  align=hdr_align, border=_HEADER_B)
    ws.row_dimensions[4].height = 22

    # Data rows
    DATA_START = 5
    shade = False

    for rec in records:
        is_sub = rec['type'] == 'subtotal'
        bg     = _YELLOW if is_sub else (_BLUE if shade else _WHITE)
        fnt    = Font(bold=True, italic=True, size=10) if is_sub else Font(size=10)
        if not is_sub:
            shade = not shade

        if include_session_col:
            vals = [rec['num'], rec['instructor'], rec['appointment_type'], rec['session'],
                    rec['subject'], rec['catalog'], rec['section'],
                    rec['course_code'], rec['title'],
                    rec['credits'], rec['teaching'], rec['pd'], rec['total']]
        else:
            vals = [rec['num'], rec['instructor'], rec['appointment_type'],
                    rec['subject'], rec['catalog'], rec['section'],
                    rec['course_code'], rec['title'],
                    rec['credits'], rec['teaching'], rec['pd'], rec['total']]

        for ci, val in enumerate(vals, 1):
            col_name = COLS[ci - 1]
            if col_name in ('#', 'Credits', 'Teaching Hours', 'PD Hours', 'Total Hours', 'Session', 'Appointment Type'):
                al = Alignment(horizontal='center', vertical='center')
            elif col_name == 'Instructor Name':
                al = Alignment(horizontal='left', vertical='center', indent=1)
            else:
                al = Alignment(horizontal='left', vertical='center')
            _set_cell(ws, DATA_START, ci, val,
                      font=fnt, fill=_fill(bg),
                      align=al, border=_THIN_B)

        ws.row_dimensions[DATA_START].height = 15
        DATA_START += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    WIDTH_MAP = {
        '#': 4, 'Instructor Name': 26, 'Appointment Type': 18, 'Session': 7,
        'Subject': 7, 'Catalog #': 8, 'Section': 7,
        'Course Code': 13, 'Class Title': 32,
        'Credits': 7, 'Teaching Hours': 10, 'PD Hours': 8, 'Total Hours': 9,
    }
    for ci, name in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WIDTH_MAP.get(name, 12)

    ws.freeze_panes = 'A5'

    buf = io.BytesIO()
    wb.save(buf)
    course_records = [r for r in records if r['type'] == 'course']
    return buf.getvalue(), course_records
