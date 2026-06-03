"""
Course Schedule Analysis — library interface
=============================================
Reads a daily schedule xlsx export and produces a five-sheet Excel workbook:

  Sheet 1 – Config                  (instructor/room overrides + room capacities)
  Sheet 2 – Capacity by Course      (enrollment vs corrected cap, transfer reserve)
  Sheet 3 – Section Detail          (every active/stop-enrl section)
  Sheet 4 – Room Availability Grid  (tracked rooms × day/time slots)
  Sheet 5 – Free Slot Summary       (standard + evening + Friday blocks)

Public API
----------
    run(schedule_path, output_path,
        existing_workbook_path=None,
        four_credit_catalogs="151,152",
        term_label="Fall 2026") -> dict
"""

import re
import warnings
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Palette & borders
# ---------------------------------------------------------------------------
C = {
    "hdr_dark":  "1F3864",
    "hdr_mid":   "2F5496",
    "hdr_light": "BDD7EE",
    "active":    "C6EFCE",
    "stop":      "FFEB9C",
    "tentative": "FCE4D6",
    "red":       "FF0000",
    "yellow":    "FF9900",
    "green":     "00B050",
    "alt":       "F2F2F2",
    "white":     "FFFFFF",
    "grid_used": "D9E1F2",
    "grid_free": "E2EFDA",
    "grid_hdr":  "4472C4",
    "grid_day":  "8EA9C1",
    "room_hdr":  "2E75B6",
}

THIN  = Side(border_style="thin",   color="BFBFBF")
THICK = Side(border_style="medium", color="595959")


def _hdr_cell(ws, row, col, value, bg=None, bold=True, color="FFFFFF",
              size=11, wrap=True, halign="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color, size=size)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    return c


def _data_cell(ws, row, col, value, bg=None, halign="left", wrap=False,
               bold=False, font_color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    if bold or font_color != "000000":
        c.font = Font(bold=bold, color=font_color)
    return c


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DEFAULT_ROOM_CAPS = {
    "NB-6.61":    26,
    "NB-6.67":    28,
    "NB-6.6402":  28,
    "NB-L2.79":   28,
    "NB-6.6337":  28,
    "NB-L2.7205": 28,
    "NB-L2.7206": 28,
    "NB-L2.7207": 28,
}

# Courses where 8 seats/section are reserved for transfer students
TRANSFER_RESERVE_COURSES = {
    ("CSCI", "373"), ("CSCI", "374"), ("CSCI", "375"), ("CSCI", "377"),
    ("MAT",  "301"),
}
TRANSFER_SEATS = 8

VALID_STATUSES = {"Active", "Stop Enrl", "Tentative"}

# Standard M/Tu/W/Th daytime + evening time blocks
STANDARD_SLOTS = [
    ("08:00", "09:15"),
    ("09:25", "10:40"),
    ("10:50", "12:05"),
    ("12:15", "13:30"),
    ("13:40", "14:55"),
    ("15:05", "16:20"),
    ("16:30", "17:45"),
    ("17:55", "20:35"),
]

STANDARD_DAYS = ["M", "Tu", "W", "Th"]

FRIDAY_BLOCKS = [
    ("08:00", "10:40"),
    ("10:50", "13:30"),
    ("13:40", "16:20"),
    ("16:30", "19:30"),
]

GRID_DAYS = ["M", "Tu", "W", "Th", "F"]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_days(days_str):
    """Parse CUNY-style days string. 'TuTh' → ['Tu','Th'], 'MW' → ['M','W']."""
    s = str(days_str).strip()
    if s in ("", "nan", "NaN", "—"):
        return []
    result, i = [], 0
    while i < len(s):
        two = s[i:i+2]
        if two in ("Tu", "Th", "Sa", "Su"):
            result.append(two); i += 2
        elif s[i] in ("M", "W", "F"):
            result.append(s[i]); i += 1
        else:
            i += 1
    return result


def _normalise_time(t_str):
    """'08:00:00' → '08:00', '8:00' → '08:00'."""
    t = str(t_str).strip()
    if t in ("", "nan", "NaN", "—"):
        return ""
    parts = t.split(":")
    try:
        return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
    except (ValueError, IndexError):
        return ""


def _fmt_time(t_str):
    """'08:00:00' or '08:00' → '8:00'."""
    t = _normalise_time(t_str)
    if not t:
        return "—"
    h, m = int(t.split(":")[0]), int(t.split(":")[1])
    return f"{h}:{m:02d}"


def _fmt_time_range(start, end):
    s, e = _fmt_time(start), _fmt_time(end)
    return "—" if s == "—" or e == "—" else f"{s}–{e}"


def _time_to_min(t_str):
    """'08:00' or '08:00:00' → minutes since midnight."""
    n = _normalise_time(t_str)
    if not n:
        return None
    h, m = int(n.split(":")[0]), int(n.split(":")[1])
    return h * 60 + m


# ---------------------------------------------------------------------------
# Schedule reader
# ---------------------------------------------------------------------------

def _read_schedule(path):
    df = pd.read_excel(path, header=1, dtype=str)

    # De-dupe columns (handles two 'Tot Enrl' columns)
    cols, seen, new_cols = list(df.columns), {}, []
    for c in cols:
        sc = str(c).strip()
        seen[sc] = seen.get(sc, -1) + 1
        new_cols.append(f"{sc}.{seen[sc]}" if seen[sc] > 0 else sc)
    df.columns = new_cols

    for col in ("Enrl Cap", "Tot Enrl"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    df["Class Stat"] = df["Class Stat"].astype(str).str.strip()
    df = df[df["Class Stat"].isin(VALID_STATUSES)].copy()

    for col in ("Subject", "Catalog#", "Section", "Class Title", "Room", "Name", "Days",
                "Mtg Start", "Mtg End"):
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().replace("nan", "")

    if "Class#" in df.columns:
        df["Class#"] = df["Class#"].astype(str).str.strip()

    return df


# ---------------------------------------------------------------------------
# Config / overrides
# ---------------------------------------------------------------------------

def _load_config(existing_wb_path, default_room_caps):
    instr_ov = {}
    room_ov  = {}
    room_caps = dict(default_room_caps)

    if not existing_wb_path or not Path(existing_wb_path).exists():
        return instr_ov, room_ov, room_caps

    try:
        wb = load_workbook(existing_wb_path, read_only=True, data_only=True)
        if "Config" not in wb.sheetnames:
            wb.close()
            return instr_ov, room_ov, room_caps

        ws   = wb["Config"]
        rows = list(ws.iter_rows(values_only=True))
        in_room_caps = False

        for row in rows:
            if not row:
                continue
            cell0 = str(row[0]).strip() if row[0] is not None else ""

            # Detect ROOM CAPACITIES section
            if cell0.upper() == "ROOM CAPACITIES":
                in_room_caps = True
                continue

            if in_room_caps:
                if not cell0 or cell0.lower() in ("room", ""):
                    continue
                try:
                    cap = int(row[1]) if len(row) > 1 and row[1] is not None else None
                    if cell0 and cap:
                        room_caps[cell0] = cap
                except (ValueError, TypeError):
                    pass
                continue

            # Instructor override (cols A–D)
            instr = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            if cell0 and instr:
                try:
                    int(cell0)
                    instr_ov[cell0] = instr
                except ValueError:
                    pass

            # Room override (cols F–I)
            room_cls = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            new_room = str(row[8]).strip() if len(row) > 8 and row[8] else ""
            if room_cls and new_room:
                try:
                    int(room_cls)
                    room_ov[room_cls] = new_room
                except ValueError:
                    pass

        wb.close()
    except Exception:
        pass

    return instr_ov, room_ov, room_caps


def _apply_overrides(df, instr_ov, room_ov):
    if not instr_ov and not room_ov:
        return df
    df = df.copy()
    for cls_num, instr in instr_ov.items():
        df.loc[df["Class#"] == cls_num, "Name"] = instr
    for cls_num, room in room_ov.items():
        df.loc[df["Class#"] == cls_num, "Room"] = room
    return df


def _add_derived_fields(df, room_caps):
    """Attach Room_Cap, Corrected_Cap, Transfer_Res to every row."""
    df = df.copy()
    room_cap_vals, corr_cap_vals, trans_res_vals = [], [], []

    for _, row in df.iterrows():
        enrl_cap = int(row.get("Enrl Cap", 0) or 0)
        room     = str(row.get("Room", "") or "")
        rc       = room_caps.get(room)
        room_cap_vals.append(rc)
        corr_cap_vals.append(enrl_cap)
        key = (str(row.get("Subject", "")).strip(), str(row.get("Catalog#", "")).strip())
        trans_res_vals.append(TRANSFER_SEATS if key in TRANSFER_RESERVE_COURSES else 0)

    df["Room_Cap"]      = room_cap_vals
    df["Corrected_Cap"] = corr_cap_vals
    df["Transfer_Res"]  = trans_res_vals
    return df


# ---------------------------------------------------------------------------
# Sheet — Capacity by Course
# ---------------------------------------------------------------------------

def _build_capacity_sheet(ws, df, term_label, schedule_name):
    ws.sheet_view.showGridLines = False
    NC = 10

    ws.merge_cells(f"A1:{get_column_letter(NC)}1")
    _hdr_cell(ws, 1, 1, f"{term_label} — Full Capacity Analysis by Course",
              bg=C["hdr_dark"], size=13, halign="left")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{get_column_letter(NC)}2")
    c = ws.cell(row=2, column=1,
                value=("8 seats/section reserved for transfers in: "
                       "CSCI 373,374,375,377,MAT 301. "
                       "Caps corrected to room capacity. "
                       "Color: red=0, yellow≤5, green>5."))
    c.font = Font(italic=True, size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 24

    hdrs = ["Course", "Course Title", "Sections",
            "Total\nCorrected Cap", "Transfer\nReserve", "Effective\nCap",
            "Total\nEnrolled", "Available\nSeats", "Fill\nRate", "Status"]
    for col, h in enumerate(hdrs, 1):
        _hdr_cell(ws, 3, col, h, bg=C["hdr_mid"])
    ws.row_dimensions[3].height = 36

    # Only Active + Stop Enrl for capacity calculations
    cap_df = df[df["Class Stat"].isin({"Active", "Stop Enrl"})].copy()

    grp = (cap_df
           .groupby(["Subject", "Catalog#", "Class Title"], sort=True)
           .agg(Sections=("Section", "count"),
                Total_Corr=("Corrected_Cap", "sum"),
                Transfer=("Transfer_Res", "sum"),
                Enrolled=("Tot Enrl", "sum"))
           .reset_index())

    grp["Course"]      = grp["Subject"] + " " + grp["Catalog#"]
    grp["Effective"]   = grp["Total_Corr"] - grp["Transfer"]

    def _status(r):
        avail = int(row["Total_Corr"]) - int(row["Transfer"]) - int(row["Enrolled"])
        effective = int(row["Total_Corr"]) - int(row["Transfer"])
        st = "Full" if avail <= 0 else "Filling" if (effective > 0 and int(row["Enrolled"]) / effective >= 0.75) else "Open"
        return st

    status_cfg = {
        "Full":    (C["red"],    "FFFFFF"),
        "Filling": (C["yellow"], "000000"),
        "Open":    (C["green"],  "FFFFFF"),
    }

    for ri, (_, row) in enumerate(grp.iterrows(), 4):
        alt   = C["alt"] if ri % 2 == 0 else C["white"]
        avail = int(row["Total_Corr"]) - int(row["Transfer"]) - int(row["Enrolled"])
        a_bg  = C["red"] if avail <= 0 else C["yellow"] if avail <= 5 else C["green"]
        a_fg  = "FFFFFF" if avail <= 0 else "000000" if avail <= 5 else "FFFFFF"
        st    = _status(row)
        st_bg, st_fg = status_cfg[st]

        _data_cell(ws, ri, 1, row["Course"],          bg=alt, bold=True)
        _data_cell(ws, ri, 2, row["Class Title"],     bg=alt)
        _data_cell(ws, ri, 3, int(row["Sections"]),   bg=alt, halign="center")
        _data_cell(ws, ri, 4, int(row["Total_Corr"]), bg=alt, halign="center")
        _data_cell(ws, ri, 5, int(row["Transfer"]),   bg=alt, halign="center")
        _data_cell(ws, ri, 6, f"=D{ri}-E{ri}",  bg=alt, halign="center")
        _data_cell(ws, ri, 7, int(row["Enrolled"]),   bg=alt, halign="center")

        c = ws.cell(row=ri, column=8, value=f"=D{ri}-E{ri}-G{ri}")
        c.fill = PatternFill("solid", start_color=a_bg)
        c.font = Font(bold=True, color=a_fg)
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row=ri, column=9, value=f"=IF(F{ri}=0,0,G{ri}/F{ri})")
        c.fill = PatternFill("solid", start_color=alt)
        c.number_format = "0.0%"
        c.alignment = Alignment(horizontal="center", vertical="center")
        c = ws.cell(row=ri, column=10, value=f'=IF(H{ri}<=0,"Full",IF(G{ri}/F{ri}>=0.75,"Filling","Open"))')
        c.fill = PatternFill("solid", start_color=st_bg)
        c.font = Font(bold=True, color=st_fg)
        c.alignment = Alignment(horizontal="center", vertical="center")

    _set_col_widths(ws, [12, 30, 9, 13, 10, 11, 11, 11, 9, 9])
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Sheet — Section Detail
# ---------------------------------------------------------------------------

def _build_section_detail_sheet(ws, df, term_label, schedule_name):
    ws.sheet_view.showGridLines = False
    NC = 14

    ws.merge_cells(f"A1:{get_column_letter(NC)}1")
    _hdr_cell(ws, 1, 1, f"{term_label} — Section-Level Capacity Detail",
              bg=C["hdr_dark"], size=13, halign="left")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{get_column_letter(NC)}2")
    c = ws.cell(row=2, column=1,
                value=f"Source: {schedule_name}  |  Active and Stop Enrl sections. "
                      "Corrected caps reflect room capacity. Overrides applied from Config sheet.")
    c.font = Font(italic=True, size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    hdrs = ["Course", "Section", "Class#", "Status", "Instructor",
            "Days", "Time", "Room",
            "Room Cap", "Enrl Cap\n(orig)", "Corrected\nCap",
            "Transfer\nRes.", "Enrolled", "Available"]
    for col, h in enumerate(hdrs, 1):
        _hdr_cell(ws, 3, col, h, bg=C["hdr_mid"])
    ws.row_dimensions[3].height = 36

    stat_color = {"Active": C["active"], "Stop Enrl": C["stop"], "Tentative": C["tentative"]}
    detail_df  = df[df["Class Stat"].isin({"Active", "Stop Enrl"})].copy()

    for ri, (_, row) in enumerate(detail_df.sort_values(["Subject", "Catalog#", "Section"]).iterrows(), 4):
        bg        = stat_color.get(row["Class Stat"], C["white"])
        enrl_cap  = int(row.get("Enrl Cap",       0) or 0)
        enrolled  = int(row.get("Tot Enrl",        0) or 0)
        corr_cap  = int(row.get("Corrected_Cap", enrl_cap))
        trans_res = int(row.get("Transfer_Res",    0))
        room_cap  = row.get("Room_Cap")
        available = corr_cap - trans_res - enrolled

        days_raw = row.get("Days", "") or ""
        days_val = days_raw if days_raw not in ("", "nan", "NaN") else "—"
        time_val = _fmt_time_range(row.get("Mtg Start", ""), row.get("Mtg End", ""))
        room_val = (row.get("Room", "") or "") or "—"
        try:
            rc_val = int(room_cap) if room_cap is not None else "—"
        except (ValueError, TypeError):
            rc_val = "—"

        _data_cell(ws, ri,  1, f"{row['Subject']} {row['Catalog#']}", bg=bg, bold=True)
        _data_cell(ws, ri,  2, row["Section"],                bg=bg, halign="center")
        _data_cell(ws, ri,  3, row.get("Class#", ""),         bg=bg, halign="center")
        _data_cell(ws, ri,  4, row["Class Stat"],             bg=bg, halign="center")
        _data_cell(ws, ri,  5, row.get("Name", ""),           bg=bg)
        _data_cell(ws, ri,  6, days_val,                      bg=bg, halign="center")
        _data_cell(ws, ri,  7, time_val,                      bg=bg, halign="center")
        _data_cell(ws, ri,  8, room_val,                      bg=bg)
        _data_cell(ws, ri,  9, rc_val,                        bg=bg, halign="center")
        _data_cell(ws, ri, 10, enrl_cap,                      bg=bg, halign="center")
        _data_cell(ws, ri, 11, corr_cap,                      bg=bg, halign="center")
        _data_cell(ws, ri, 12, trans_res,                     bg=bg, halign="center")
        _data_cell(ws, ri, 13, enrolled,                      bg=bg, halign="center")
        _data_cell(ws, ri, 14, available,                     bg=bg, halign="center")

    _set_col_widths(ws, [12, 8, 9, 11, 30, 7, 13, 14, 9, 10, 10, 9, 9, 10])
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Sheet — Room Availability Grid
# ---------------------------------------------------------------------------

def _overlaps(sec_start_min, sec_end_min, slot_start_min, slot_end_min):
    """True if section [sec_start, sec_end) overlaps slot [slot_start, slot_end)."""
    return sec_start_min < slot_end_min and sec_end_min > slot_start_min


def _build_room_grid_sheet(ws, df, room_caps, term_label, schedule_name):
    ws.sheet_view.showGridLines = False

    tracked = sorted(r for r in room_caps if not r.upper().startswith("OL"))
    if not tracked:
        ws.cell(row=1, column=1, value="No tracked rooms configured.")
        return

    def short(r):
        return r.replace("NB-", "").replace("NB ", "")

    # Use Active + Stop Enrl — both physically occupy the room
    sched_df = df[
        (df["Class Stat"].isin({"Active", "Stop Enrl"})) &
        (df["Room"].isin(tracked)) &
        (df["Mtg Start"].str.len() > 0)
    ].copy()

    if sched_df.empty:
        ws.cell(row=1, column=1, value="No timetable data found for tracked rooms.")
        return

    # Use fixed STANDARD_SLOTS as grid rows — overlap-based matching so that
    # non-standard blocks (e.g. a 10:50–13:30 Friday class) correctly mark
    # both the 10:50–12:05 and 12:15–13:30 standard rows as occupied.
    slots = STANDARD_SLOTS   # list of ("HH:MM", "HH:MM") tuples

    # Precompute section info keyed by (room, day) for fast lookup
    # Each entry: list of (sec_start_min, sec_end_min, text)
    room_day_sections: dict = {}
    for _, row in sched_df.iterrows():
        s_norm = _normalise_time(row["Mtg Start"])
        e_norm = _normalise_time(row["Mtg End"])
        if not s_norm or not e_norm:
            continue
        room   = row["Room"]
        course = f"{row['Subject']} {row['Catalog#']}"
        sect   = row["Section"]
        enrl   = int(row.get("Tot Enrl", 0) or 0)
        corr   = int(row.get("Corrected_Cap", 0) or 0)
        text   = f"{course}\n§{sect}\n{enrl}/{corr}"
        sm     = _time_to_min(s_norm)
        em     = _time_to_min(e_norm)
        for d in _parse_days(row.get("Days", "")):
            if d in GRID_DAYS and sm is not None and em is not None:
                room_day_sections.setdefault((room, d), []).append((sm, em, text))

    def _cell_for(room, day, slot_s, slot_e):
        """Return section text if any section overlaps this standard slot, else None."""
        sm = _time_to_min(slot_s)
        em = _time_to_min(slot_e)
        for (sec_sm, sec_em, text) in room_day_sections.get((room, day), []):
            if _overlaps(sec_sm, sec_em, sm, em):
                return text
        return None

    ND   = len(GRID_DAYS)
    NC   = 1 + len(tracked) * ND
    rooms_lbl = " · ".join(short(r) for r in tracked)

    # Row 1 — title
    ws.merge_cells(f"A1:{get_column_letter(NC)}1")
    _hdr_cell(ws, 1, 1, f"{term_label} — Room Availability Grid ({rooms_lbl})",
              bg=C["hdr_dark"], size=12, halign="left")
    ws.row_dimensions[1].height = 26

    # Row 2 — subtitle
    ws.merge_cells(f"A2:{get_column_letter(NC)}2")
    c = ws.cell(row=2, column=1,
                value=f"Source: {schedule_name}  |  Active sections only. Config overrides applied.")
    c.font = Font(italic=True, size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Row 3 — "Time Slot" + room headers (each spanning ND cols)
    c = ws.cell(row=3, column=1, value="Time Slot")
    c.font  = Font(bold=True, color="FFFFFF", size=10)
    c.fill  = PatternFill("solid", start_color=C["grid_hdr"])
    c.alignment = Alignment(horizontal="center", vertical="center")

    for ri, room in enumerate(tracked):
        sc  = 2 + ri * ND
        ec  = sc + ND - 1
        ws.merge_cells(start_row=3, start_column=sc, end_row=3, end_column=ec)
        cap = room_caps.get(room, "?")
        c   = ws.cell(row=3, column=sc, value=f"{short(room)}  (cap {cap})")
        c.font  = Font(bold=True, color="FFFFFF", size=10)
        c.fill  = PatternFill("solid", start_color=C["room_hdr"])
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    # Row 4 — day labels
    c = ws.cell(row=4, column=1)
    c.fill = PatternFill("solid", start_color=C["grid_hdr"])
    for ri, room in enumerate(tracked):
        for di, day in enumerate(GRID_DAYS):
            col = 2 + ri * ND + di
            c   = ws.cell(row=4, column=col, value=day)
            c.font  = Font(bold=True, color="FFFFFF", size=10)
            c.fill  = PatternFill("solid", start_color=C["grid_day"])
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 18

    # Row 5 — empty separator
    ws.row_dimensions[5].height = 4

    # Data rows — one per standard time slot
    for si, (s, e) in enumerate(slots):
        rw  = 6 + si
        lbl = f"{_fmt_time(s)}–{_fmt_time(e)}"

        c = ws.cell(row=rw, column=1, value=lbl)
        c.font  = Font(bold=True, size=9)
        c.fill  = PatternFill("solid", start_color=C["hdr_light"])
        c.alignment = Alignment(horizontal="center", vertical="center")

        for ri, room in enumerate(tracked):
            for di, day in enumerate(GRID_DAYS):
                col  = 2 + ri * ND + di
                text = _cell_for(room, day, s, e)
                bg   = C["grid_used"] if text else C["grid_free"]
                val  = text if text else "FREE"
                c    = ws.cell(row=rw, column=col, value=val)
                c.fill = PatternFill("solid", start_color=bg)
                c.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
                c.font = Font(size=8, bold=bool(text),
                              color="1F3864" if text else "375623")
        ws.row_dimensions[rw].height = 42

    # Column widths
    ws.column_dimensions["A"].width = 13
    for ri in range(len(tracked)):
        for di in range(ND):
            ws.column_dimensions[get_column_letter(2 + ri * ND + di)].width = 11

    ws.freeze_panes = "B6"


# ---------------------------------------------------------------------------
# Sheet — Free Slot Summary
# ---------------------------------------------------------------------------

def _build_free_slots_sheet(ws, df, room_caps, term_label):
    ws.sheet_view.showGridLines = False

    tracked   = sorted(r for r in room_caps if not r.upper().startswith("OL"))
    rooms_lbl = ", ".join(tracked)

    ws.merge_cells("A1:E1")
    _hdr_cell(ws, 1, 1, f"{term_label} — Free Room Time Slots ({rooms_lbl})",
              bg=C["hdr_dark"], size=12, halign="left")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:E2")
    c = ws.cell(row=2, column=1,
                value="Standard M/Tu/W/Th time blocks and Friday extended blocks "
                      "with no scheduled section.")
    c.font = Font(italic=True, size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    hdrs = ["Room", "Day", "Block Start", "Block End", "Notes"]
    for col, h in enumerate(hdrs, 1):
        _hdr_cell(ws, 3, col, h, bg=C["hdr_mid"])
    ws.row_dimensions[3].height = 22

    # Active + Stop Enrl both physically occupy the room
    sched_df = df[
        df["Class Stat"].isin({"Active", "Stop Enrl"}) &
        df["Room"].isin(tracked)
    ]

    # Build per-(room, day) list of (start_min, end_min) for overlap checks
    room_day_mins: dict = {}
    for _, row in sched_df.iterrows():
        room = row.get("Room", "") or ""
        s = _normalise_time(row.get("Mtg Start", ""))
        e = _normalise_time(row.get("Mtg End", ""))
        if not s or not e:
            continue
        sm, em = _time_to_min(s), _time_to_min(e)
        if sm is None or em is None:
            continue
        for d in _parse_days(row.get("Days", "")):
            room_day_mins.setdefault((room, d), []).append((sm, em))

    def _is_blocked(room, day, slot_s, slot_e):
        """True if any section overlaps the given slot for this room+day."""
        sm = _time_to_min(slot_s)
        em = _time_to_min(slot_e)
        for (sec_sm, sec_em) in room_day_mins.get((room, day), []):
            if _overlaps(sec_sm, sec_em, sm, em):
                return True
        return False

    def _note(day, start):
        if day == "F":
            return "Friday extended"
        m = _time_to_min(start) or 0
        if m >= 17 * 60 + 55:
            return "Evening"
        return "Standard slot"

    rw = 4
    for room in tracked:
        # M/Tu/W/Th — check every standard slot
        for day in STANDARD_DAYS:
            for (s, e) in STANDARD_SLOTS:
                if not _is_blocked(room, day, s, e):
                    alt = C["alt"] if rw % 2 == 0 else C["white"]
                    _data_cell(ws, rw, 1, room,          bg=alt)
                    _data_cell(ws, rw, 2, day,           bg=alt, halign="center")
                    _data_cell(ws, rw, 3, _fmt_time(s),  bg=alt, halign="center")
                    _data_cell(ws, rw, 4, _fmt_time(e),  bg=alt, halign="center")
                    _data_cell(ws, rw, 5, _note(day, s), bg=alt)
                    rw += 1
        # Friday extended blocks — free only if nothing overlaps
        for (s, e) in FRIDAY_BLOCKS:
            if not _is_blocked(room, "F", s, e):
                alt = C["alt"] if rw % 2 == 0 else C["white"]
                _data_cell(ws, rw, 1, room,              bg=alt)
                _data_cell(ws, rw, 2, "F",               bg=alt, halign="center")
                _data_cell(ws, rw, 3, _fmt_time(s),      bg=alt, halign="center")
                _data_cell(ws, rw, 4, _fmt_time(e),      bg=alt, halign="center")
                _data_cell(ws, rw, 5, "Friday extended",  bg=alt)
                rw += 1

    _set_col_widths(ws, [16, 7, 12, 12, 18])
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Config sheet
# ---------------------------------------------------------------------------

def _ensure_config_sheet(wb, term_label, room_caps):
    if "Config" in wb.sheetnames:
        return
    ws = wb.create_sheet("Config")
    ws.sheet_view.showGridLines = False

    # Row 1: title
    ws.merge_cells("A1:I1")
    _hdr_cell(ws, 1, 1,
              f"{term_label} — Configuration  (instructor overrides & room adjustments)",
              bg=C["hdr_mid"], size=12, halign="left")
    ws.row_dimensions[1].height = 24

    # Row 2: instructions
    ws.merge_cells("A2:I2")
    c = ws.cell(row=2, column=1,
                value=("Edit these tables to override instructor names or room assignments. "
                       "Changes take effect on next refresh."))
    c.font = Font(italic=True, size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6  # spacer

    # Row 4: section headers
    ws.merge_cells("A4:D4")
    _hdr_cell(ws, 4, 1, "INSTRUCTOR OVERRIDES", bg=C["hdr_light"],
              color="000000", size=10)
    ws.merge_cells("F4:I4")
    _hdr_cell(ws, 4, 6, "ROOM ADJUSTMENTS", bg=C["hdr_light"],
              color="000000", size=10)

    # Row 5: column headers
    for col, h in enumerate(["Class#", "Course", "Section", "Instructor Name"], 1):
        _hdr_cell(ws, 5, col, h, bg=C["hdr_light"], color="000000", size=10)
    for col, h in enumerate(["Class#", "Course", "Section", "New Room"], 6):
        _hdr_cell(ws, 5, col, h, bg=C["hdr_light"], color="000000", size=10)

    for r in range(6, 17):
        ws.row_dimensions[r].height = 15

    ws.row_dimensions[17].height = 8  # spacer

    # Row 18: ROOM CAPACITIES header
    ws.merge_cells("A18:B18")
    _hdr_cell(ws, 18, 1, "ROOM CAPACITIES", bg=C["hdr_light"],
              color="000000", size=10)
    ws.row_dimensions[18].height = 18

    # Row 19: column headers
    _hdr_cell(ws, 19, 1, "Room",     bg=C["hdr_light"], color="000000", size=10)
    _hdr_cell(ws, 19, 2, "Capacity", bg=C["hdr_light"], color="000000", size=10)

    # Populate with any rooms already configured — empty on first run
    for idx, (room, cap) in enumerate(sorted(room_caps.items()), 20):
        ws.cell(row=idx, column=1, value=room).alignment = Alignment(vertical="center")
        ws.cell(row=idx, column=2, value=cap).alignment  = Alignment(
            horizontal="center", vertical="center")
        ws.row_dimensions[idx].height = 15

    note_row = 20 + len(room_caps)
    c = ws.cell(row=note_row, column=1,
                value="← Add rooms here or use the Manage Rooms button in the dashboard.")
    c.font = Font(italic=True, size=9, color="595959")
    ws.merge_cells(f"A{note_row}:D{note_row}")

    _set_col_widths(ws, [10, 12, 10, 30, 4, 10, 12, 10, 20])


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def run(
    schedule_path: str,
    output_path: str,
    existing_workbook_path: str = None,
    four_credit_catalogs: str = "151,152",
    term_label: str = "Fall 2026",
    room_config_path: str = None,
) -> dict:
    """
    Run the full schedule analysis and write an Excel workbook.

    Parameters
    ----------
    schedule_path          : path to the schedule xlsx export
    output_path            : where to write the output workbook
    existing_workbook_path : if supplied, Config sheet overrides are preserved
    four_credit_catalogs   : comma-separated catalog numbers for 4-credit courses
    term_label             : label used in all sheet titles
    room_config_path       : optional path to room_config.json set via the Manage Rooms UI

    Returns dict with summary counts and output_path.
    """
    run_date      = datetime.today().strftime("%Y-%m-%d %H:%M")
    schedule_name = Path(schedule_path).name

    df = _read_schedule(schedule_path)

    # Start with no default rooms — users configure their rooms via the UI
    instr_ov, room_ov, room_caps = _load_config(existing_workbook_path, {})

    # Merge UI-configured rooms (set via the Manage Rooms UI)
    if room_config_path and Path(room_config_path).exists():
        try:
            import json as _jmod
            for r in _jmod.loads(Path(room_config_path).read_text()):
                if isinstance(r, dict) and r.get('name') and r.get('capacity'):
                    room_caps[r['name']] = int(r['capacity'])
        except Exception:
            pass

    df = _apply_overrides(df, instr_ov, room_ov)
    df = _add_derived_fields(df, room_caps)

    active    = int((df["Class Stat"] == "Active").sum())
    stop_enrl = int((df["Class Stat"] == "Stop Enrl").sum())
    tentative = int((df["Class Stat"] == "Tentative").sum())

    # Start from existing workbook (preserves Config) or fresh
    if existing_workbook_path and Path(existing_workbook_path).exists():
        wb = load_workbook(existing_workbook_path)
        for name in ("Capacity by Course", "Section Detail",
                     "Room Availability Grid", "Free Slot Summary"):
            if name in wb.sheetnames:
                del wb[name]
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    _ensure_config_sheet(wb, term_label, room_caps)

    ci  = wb.sheetnames.index("Config")
    ws1 = wb.create_sheet("Capacity by Course",     ci + 1)
    ws2 = wb.create_sheet("Section Detail",         ci + 2)
    ws3 = wb.create_sheet("Room Availability Grid", ci + 3)
    ws4 = wb.create_sheet("Free Slot Summary",      ci + 4)

    _build_capacity_sheet(ws1, df, term_label, schedule_name)
    _build_section_detail_sheet(ws2, df, term_label, schedule_name)
    _build_room_grid_sheet(ws3, df, room_caps, term_label, schedule_name)
    _build_free_slots_sheet(ws4, df, room_caps, term_label)

    wb.save(output_path)

    return {
        "active":            active,
        "stop_enrl":         stop_enrl,
        "tentative":         tentative,
        "total":             active + stop_enrl + tentative,
        "rooms":             int(df["Room"].nunique()),
        "courses":           int(df.groupby(["Subject", "Catalog#"]).ngroups),
        "overrides_applied": len(instr_ov) + len(room_ov),
        "output_path":       output_path,
        "run_date":          run_date,
    }
