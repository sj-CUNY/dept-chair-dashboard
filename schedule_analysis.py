"""
Fall 2026 Full Analysis — Linked Workbook Builder
Usage:  python build_linked.py <schedule_file.xlsx>
If no argument given, uses the most recently modified schedule xlsx in the workspace.
Reads Config sheet from the existing workbook (instructor overrides + room adjustments).
Rewrites all analysis sheets in-place, preserving Config.
"""
import sys, os, glob, pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import time as dtime

WORKBOOK_PATH = 'Fall2026_FullAnalysis.xlsx'
WORKSPACE     = '.'

# ── RESOLVE SCHEDULE FILE ──────────────────────────────────────────────────────
if len(sys.argv) > 1:
    sched_path = sys.argv[1]
else:
    # Auto-pick most recently modified schedule xlsx in workspace
    candidates = [f for f in glob.glob(WORKSPACE + '*.xlsx')
                  if os.path.basename(f) != 'Fall2026_FullAnalysis.xlsx'
                  and os.path.basename(f) != 'Fall2026_RoomChanges.xlsx']
    if not candidates:
        sys.exit('No schedule file found. Pass path as argument.')
    sched_path = max(candidates, key=os.path.getmtime)

print(f'Schedule file : {os.path.basename(sched_path)}')

# ── LOAD SCHEDULE ──────────────────────────────────────────────────────────────
sched     = pd.read_excel(sched_path, header=1)
active    = sched[sched['Class Stat'] == 'Active'].copy().reset_index(drop=True)
stop_enrl = sched[sched['Class Stat'] == 'Stop Enrl'].copy().reset_index(drop=True)

# ── READ CONFIG FROM EXISTING WORKBOOK ────────────────────────────────────────
instr_overrides = {}   # Class# → name
room_adjustments = {}  # Class# → room

if os.path.exists(WORKBOOK_PATH):
    try:
        cfg_wb = load_workbook(WORKBOOK_PATH, data_only=True)
        if 'Config' in cfg_wb.sheetnames:
            ws_cfg = cfg_wb['Config']
            # Instructor overrides table starts at A5
            for row in ws_cfg.iter_rows(min_row=6, values_only=True):
                if row[0] and str(row[0]).strip() not in ('', 'Class#'):
                    try:
                        cn = int(row[0])
                        if row[3]: instr_overrides[cn] = str(row[3]).strip()
                    except: pass
            # Room adjustments table starts at F5
            for row in ws_cfg.iter_rows(min_row=6, values_only=True):
                if row[5] and str(row[5]).strip() not in ('', 'Class#'):
                    try:
                        cn = int(row[5])
                        if row[8]: room_adjustments[cn] = str(row[8]).strip()
                    except: pass
        cfg_wb.close()
    except Exception as e:
        print(f'  Config read warning: {e}')

print(f'Instructor overrides: {len(instr_overrides)}')
print(f'Room adjustments    : {len(room_adjustments)}')

# Apply overrides
for cn, name in instr_overrides.items():
    active.loc[active['Class#'] == cn, 'Name'] = name
for cn, room in room_adjustments.items():
    active.loc[active['Class#'] == cn, 'Room'] = room

# ── CONSTANTS ──────────────────────────────────────────────────────────────────
ROOM_CAPS = {'NB-6.61':26,'NB-6.67':28,'NB-6.6402':28,
             'NB-L2.79':28,'NB-L2.7205':28,'NB-L2.7206':28,'NB-L2.7207':28}
TRACKED_ROOMS   = ['NB-6.61','NB-6.67','NB-6.6402','NB-L2.79']
TRANSFER_COURSES= {('CSCI',373),('CSCI',374),('CSCI',375),('CSCI',377),('MAT',301)}

def room_cap(r):
    if pd.isna(r) or str(r).startswith('OL'): return None
    return ROOM_CAPS.get(str(r))

def corr_cap(row):
    if row['Subject']=='FCM': return row['Enrl Cap']
    c = room_cap(row['Room'])
    return c if c else row['Enrl Cap']

def trans_res(row):
    return 8 if (row['Subject'],row['Catalog#']) in TRANSFER_COURSES else 0

for df in [active, stop_enrl]:
    df['RoomCap']  = df['Room'].apply(room_cap)
    df['CorrCap']  = df.apply(corr_cap, axis=1)
    df['TransRes'] = df.apply(trans_res, axis=1)
    df['EffCap']   = df['CorrCap'] - df['TransRes']
    df['Available']= df['EffCap'] - df['Tot Enrl']
    df['Course']   = df['Subject'] + ' ' + df['Catalog#'].astype(str)

# ── STYLE HELPERS ──────────────────────────────────────────────────────────────
def fill(h): return PatternFill('solid',fgColor=h)
def font(h='FF000000',bold=False,size=10,name='Arial'):
    return Font(color=h,bold=bold,size=size,name=name)
def center(): return Alignment(horizontal='center',vertical='center',wrap_text=True)
def left():   return Alignment(horizontal='left',  vertical='center',wrap_text=True)

FT=fill('FFD6E4F7'); FN=fill('FFEBF3FB'); FH=fill('FF1F3864')
FA=fill('FFF5F5F3'); FW=fill('FFFFFFFF'); FG=fill('FFE2EFDA')
FY=fill('FFFFF2CC'); FR=fill('FFFDECEA'); FGR=fill('FFF5F5F3')
FS=fill('FFDCE6F1')
fT=font('FF1F3864',True); fN=font('FF555555'); fH=font('FFFFFFFF',True)
fB=font(); fG=font('FF375623',True); fY=font('FF7F6000',True)
fR=font('FFA32D2D',True); fD=font('FF888888')
fST=Font(color='FF595959',size=10,name='Arial',italic=True)
fSB=Font(color='FF595959',size=10,name='Arial',italic=True,bold=True)

RSTYLE={'NB-6.61':{'hf':fill('FF2E75B6'),'cf':fill('FFD6E4F7'),'fn':font('FF2E75B6')},
        'NB-6.67':{'hf':fill('FF375623'),'cf':fill('FFE2EFDA'),'fn':font('FF375623')},
        'NB-6.6402':{'hf':fill('FFD85A30'),'cf':fill('FFFCE4D6'),'fn':font('FFD85A30')},
        'NB-L2.79':{'hf':fill('FF7030A0'),'cf':fill('FFEFE0F5'),'fn':font('FF7030A0')}}

def apc(ws,r,c,v,f,fn,al=None):
    cell=ws.cell(row=r,column=c,value=v)
    cell.fill=f; cell.font=fn; cell.alignment=al or center(); return cell

def ss(av,fr):
    if av<=0: return 'Full',FR,fR
    if av<=5: return 'Near Full',FR,fR
    if fr>=0.75: return 'Filling',FY,fY
    return 'Open',FG,fG

def avs(av):
    if av<=0: return FR,fR
    if av<=5: return FY,fY
    return FG,fG

def fmt_t(t):
    if pd.isna(t): return None
    try: return f'{t.hour}:{t.minute:02d}'
    except: return str(t)[:5]

SO = {'CSCI':0,'FCM':1,'MAT':2,'STA':3}

# ── CONFIG SHEET ───────────────────────────────────────────────────────────────
def build_config(wb):
    ws = wb.create_sheet('Config')
    ws.sheet_properties.tabColor = 'FF7030A0'

    ws.merge_cells('A1:I1')
    apc(ws,1,1,'Fall 2026 — Configuration  (instructor overrides & room adjustments)',FH,fH,left())
    ws.row_dimensions[1].height = 20

    ws.merge_cells('A2:I2')
    apc(ws,2,1,'Edit these tables to override instructor names or room assignments. Changes take effect on next refresh.',FN,fN,left())
    ws.row_dimensions[2].height = 14

    # Left table: Instructor overrides
    apc(ws,4,1,'INSTRUCTOR OVERRIDES',FH,fH,left()); ws.merge_cells('A4:D4')
    for i,h in enumerate(['Class#','Course','Section','Instructor Name'],1):
        c=ws.cell(row=5,column=i,value=h); c.fill=fill('FF2E75B6'); c.font=fH; c.alignment=center()
    ws.row_dimensions[5].height = 16

    # Right table: Room adjustments
    apc(ws,4,6,'ROOM ADJUSTMENTS',FH,fH,left()); ws.merge_cells('F4:I4')
    for i,h in enumerate(['Class#','Course','Section','New Room'],1):
        c=ws.cell(row=5,column=5+i,value=h); c.fill=fill('FFD85A30'); c.font=fH; c.alignment=center()

    # Populate instructor overrides
    io_data = [(cn, '', '', nm) for cn, nm in sorted(instr_overrides.items())]
    for ri, (cn, crs, sec, nm) in enumerate(io_data, start=6):
        row_data = active[active['Class#']==cn]
        if not row_data.empty:
            r = row_data.iloc[0]
            crs=r['Course']; sec=str(r['Section'])
        rf = FA if ri%2==0 else FW
        for ci,v in enumerate([cn,crs,sec,nm],1):
            c=ws.cell(row=ri,column=ci,value=v); c.fill=rf; c.font=fB; c.alignment=center()
        ws.row_dimensions[ri].height = 14
    # Add empty rows for user to add more
    for ri in range(6+len(io_data), 6+len(io_data)+10):
        rf = FA if ri%2==0 else FW
        for ci in range(1,5):
            ws.cell(row=ri,column=ci).fill = rf
        ws.row_dimensions[ri].height = 14

    # Populate room adjustments
    ra_data = [(cn,'','',rm) for cn,rm in sorted(room_adjustments.items())]
    for ri, (cn, crs, sec, rm) in enumerate(ra_data, start=6):
        row_data = active[active['Class#']==cn]
        if not row_data.empty:
            r=row_data.iloc[0]; crs=r['Course']; sec=str(r['Section'])
        rf = FA if ri%2==0 else FW
        for ci,v in enumerate([cn,crs,sec,rm],1):
            c=ws.cell(row=ri,column=5+ci,value=v); c.fill=rf; c.font=fB; c.alignment=center()
        ws.row_dimensions[ri].height = 14
    for ri in range(6+len(ra_data), 6+len(ra_data)+10):
        rf = FA if ri%2==0 else FW
        for ci in range(1,5):
            ws.cell(row=ri,column=5+ci).fill = rf
        ws.row_dimensions[ri].height = 14

    for col,w in zip('ABCDEFGHI',[9,14,9,22,3,9,14,9,16]):
        ws.column_dimensions[col].width = w

# ── SECTION DETAIL ─────────────────────────────────────────────────────────────
def build_section_detail(wb):
    ws = wb.create_sheet('Section Detail')
    ws.merge_cells('A1:N1')
    apc(ws,1,1,'Fall 2026 — Section-Level Capacity Detail',FT,fT,left())
    ws.row_dimensions[1].height=20
    ws.merge_cells('A2:N2')
    src=os.path.basename(sched_path)
    apc(ws,2,1,f'Source: {src}  |  Active and Stop Enrl sections. Corrected caps reflect room capacity. Overrides applied from Config sheet.',FN,fN,left())
    ws.row_dimensions[2].height=15
    hdrs=['Course','Section','Class#','Status','Instructor','Days','Time','Room','Room Cap',
          'Enrl Cap\n(orig)','Corrected\nCap','Transfer\nRes.','Enrolled','Available']
    for i,h in enumerate(hdrs,1):
        c=ws.cell(row=3,column=i,value=h); c.fill=FH; c.font=fH; c.alignment=center()
    ws.row_dimensions[3].height=28
    for col,w in zip('ABCDEFGHIJKLMN',[13,9,8,10,30,8,14,12,9,10,11,10,9,10]):
        ws.column_dimensions[col].width=w

    act=active.copy(); act['_st']='Active'
    stp=stop_enrl.copy(); stp['_st']='Stop Enrl'
    combo=pd.concat([act,stp],ignore_index=True)
    combo['_so']=combo['Subject'].map(SO).fillna(9)
    combo['_s']=combo['Section'].astype(str)
    combo=combo.sort_values(['_so','Catalog#','_s'])

    row=4; prev=None; tog=True
    for _,r in combo.iterrows():
        course=r['Course']; is_stop=r['_st']=='Stop Enrl'
        if course!=prev: prev=course; tog=True
        rf=FS if is_stop else (FA if tog else FW)
        if not is_stop: tog=not tog
        instr=r['Name'] if pd.notna(r.get('Name')) else '—'
        rm=str(r['Room']) if pd.notna(r['Room']) else None
        rc=int(r['RoomCap']) if pd.notna(r['RoomCap']) else '—'
        days=str(r['Days']) if pd.notna(r['Days']) else '—'
        ts=fmt_t(r['Mtg Start']); te=fmt_t(r['Mtg End'])
        time_str=f'{ts}–{te}' if ts and te else '—'
        vals=[course,str(r['Section']),int(r['Class#']),r['_st'],instr,days,time_str,rm,rc,
              int(r['Enrl Cap']),int(r['CorrCap']),int(r['TransRes']),int(r['Tot Enrl']),int(r['Available'])]
        for ci,val in enumerate(vals,1):
            c=ws.cell(row=row,column=ci,value=val); c.fill=rf; c.alignment=center()
            c.font=fSB if (is_stop and ci==1) else fST if is_stop else Font(color='FF000000',bold=(ci==1),size=10,name='Arial')
        ws.row_dimensions[row].height=15; row+=1

# ── CAPACITY BY COURSE ─────────────────────────────────────────────────────────
def build_capacity(wb):
    ws=wb.create_sheet('Capacity by Course')
    ws.merge_cells('A1:J1')
    apc(ws,1,1,'Fall 2026 — Full Capacity Analysis by Course',FT,fT,left())
    ws.row_dimensions[1].height=20
    ws.merge_cells('A2:J2')
    apc(ws,2,1,'8 seats/section reserved for transfers in: CSCI 373,374,375,377,MAT 301. Caps corrected to room capacity. Color: red=0, yellow≤5, green>5.',FN,fN,left())
    ws.row_dimensions[2].height=15
    for i,h in enumerate(['Course','Course Title','Sections','Total\nCorrected Cap','Transfer\nReserve',
                          'Effective\nCap','Total\nEnrolled','Available\nSeats','Fill\nRate','Status'],1):
        c=ws.cell(row=3,column=i,value=h); c.fill=FH; c.font=fH; c.alignment=center()
    ws.row_dimensions[3].height=28
    for col,w in zip('ABCDEFGHIJ',[13,30,9,16,14,13,13,13,10,14]):
        ws.column_dimensions[col].width=w

    a2=active.copy(); a2['_so']=a2['Subject'].map(SO).fillna(9)
    grp=a2.groupby(['_so','Course']).agg(Title=('Class Title','first'),Sects=('Section','count'),
        TCC=('CorrCap','sum'),TTR=('TransRes','sum'),TE=('Tot Enrl','sum')).reset_index().sort_values(['_so','Course'])
    grp['Eff']=grp['TCC']-grp['TTR']; grp['Av']=grp['Eff']-grp['TE']
    grp['FR']=grp['TE']/grp['Eff'].replace(0,float('nan'))

    row=4; tog=True
    for _,r in grp.iterrows():
        rf=FA if tog else FW; tog=not tog
        av=int(r['Av']); fr=r['FR'] if pd.notna(r['FR']) else 0
        st,sf,sn=ss(av,fr); af,an=avs(av)
        fp=f"{fr*100:.1f}%" if pd.notna(r['FR']) else '0.0%'
        data=[(r['Course'],rf,Font(color='FF000000',bold=True,name='Arial',size=10),left()),
              (r['Title'],rf,fB,left()),(int(r['Sects']),rf,fB,center()),
              (int(r['TCC']),rf,fB,center()),(int(r['TTR']),rf,fB,center()),
              (int(r['Eff']),rf,fB,center()),(int(r['TE']),rf,fB,center()),
              (av,af,an,center()),(fp,rf,fB,center()),(st,sf,sn,center())]
        for ci,(v,f,fn,al) in enumerate(data,1):
            c=ws.cell(row=row,column=ci,value=v); c.fill=f; c.font=fn; c.alignment=al
        ws.row_dimensions[row].height=15; row+=1

# ── ROOM GRID ──────────────────────────────────────────────────────────────────
STD_SLOTS=[('08:00–09:15',dtime(8,0),dtime(9,15)),('09:25–10:40',dtime(9,25),dtime(10,40)),
           ('10:50–12:05',dtime(10,50),dtime(12,5)),('12:15–13:30',dtime(12,15),dtime(13,30)),
           ('15:05–16:20',dtime(15,5),dtime(16,20)),('16:30–17:45',dtime(16,30),dtime(17,45)),
           ('17:55–20:35',dtime(17,55),dtime(20,35)),('18:00–20:00',dtime(18,0),dtime(20,0))]
FRI_EXT=[('F 08:00–10:40',dtime(8,0),dtime(10,40)),('F 10:50–13:30',dtime(10,50),dtime(13,30)),
         ('F 13:40–16:20',dtime(13,40),dtime(16,20)),('F 15:05–17:45',dtime(15,5),dtime(17,45))]

def parse_days(d):
    if pd.isna(d): return []
    d=str(d).strip(); r=[]
    for x in ['TuTh','MW','Tu','Th','M','W','F']:
        if x in d:
            if x=='TuTh': r+=['Tu','Th']
            elif x=='MW': r+=['M','W']
            else: r.append(x)
            d=d.replace(x,'')
    return list(dict.fromkeys(r))

def t2time(t):
    if pd.isna(t): return None
    try: return dtime(t.hour, t.minute)
    except:
        try: total=int(t.total_seconds()); h,m=divmod(total//60,60); return dtime(h,m)
        except: return None

def build_grid(wb):
    ws=wb.create_sheet('Room Availability Grid')
    tr=active[active['Room'].isin(TRACKED_ROOMS)].copy()
    tr['ST']=tr['Mtg Start'].apply(t2time); tr['ET']=tr['Mtg End'].apply(t2time)
    tr['DL']=tr['Days'].apply(parse_days)
    def mct(r): return f"{r['Subject']} {r['Catalog#']}\n§{r['Section']}\n{int(r['Tot Enrl'])}/{int(r['EffCap'])}"
    tr['CT']=tr.apply(mct,axis=1)
    occ={}
    for _,r in tr.iterrows():
        rm=r['Room']; st=r['ST']; et=r['ET']
        if not st or not et: continue
        for day in r['DL']:
            if day=='F':
                for lbl,ss2,se in STD_SLOTS:
                    if st<=ss2 and et>ss2:
                        k=(rm,day,lbl)
                        if k not in occ: occ[k]=r['CT']
                for lbl,fs,fe in FRI_EXT:
                    if st==fs:
                        k=(rm,day,lbl)
                        if k not in occ: occ[k]=r['CT']
            else:
                for lbl,ss2,se in STD_SLOTS:
                    if st==ss2:
                        k=(rm,day,lbl)
                        if k not in occ: occ[k]=r['CT']

    src=os.path.basename(sched_path)
    ws.merge_cells('A1:W1')
    apc(ws,1,1,'Fall 2026 — Room Availability Grid (NB-6.61 · NB-6.67 · NB-6.6402 · NB-L2.79)',FT,fT,left())
    ws.row_dimensions[1].height=22
    ws.merge_cells('A2:W2')
    apc(ws,2,1,f'Source: {src}  |  Active sections only. Config overrides applied.',FN,fN,left())
    ws.row_dimensions[2].height=15
    apc(ws,3,1,'Time Slot',FH,fH,center()); ws.row_dimensions[3].height=20
    RC=[('NB-6.61',2),('NB-6.67',7),('NB-6.6402',12),('NB-L2.79',17)]
    for rm,sc in RC:
        st2=RSTYLE[rm]; cap=ROOM_CAPS[rm]
        ws.merge_cells(start_row=3,start_column=sc,end_row=3,end_column=sc+4)
        apc(ws,3,sc,f"{rm.replace('NB-','')}  (cap {cap})",st2['hf'],fH,center())
    DO=['M','Tu','W','Th','F']
    ws.cell(row=4,column=1).fill=FW; ws.row_dimensions[4].height=18
    for rm,sc in RC:
        st2=RSTYLE[rm]
        for di,day in enumerate(DO):
            c=ws.cell(row=4,column=sc+di,value=day); c.fill=st2['hf']; c.font=fH; c.alignment=center()
    for col in range(1,22): ws.cell(row=5,column=col).fill=FW
    ws.row_dimensions[5].height=4
    for gr,lbl in enumerate(['08:00–09:15','09:25–10:40','10:50–12:05','12:15–13:30','15:05–16:20','16:30–17:45'],6):
        apc(ws,gr,1,lbl,FN,Font(color='FF1F3864',bold=True,size=10,name='Arial'),center())
        for rm,sc in RC:
            st2=RSTYLE[rm]
            for di,day in enumerate(DO):
                k=(rm,day,lbl)
                if k in occ: apc(ws,gr,sc+di,occ[k],st2['cf'],st2['fn'],center())
                else:         apc(ws,gr,sc+di,'FREE',FG,fG,center())
        ws.row_dimensions[gr].height=36
    apc(ws,12,1,'13:30–15:05  (gap)',FGR,fD,center())
    for col in range(2,22): ws.cell(row=12,column=col).fill=FGR
    ws.row_dimensions[12].height=11
    apc(ws,13,1,'Friday Extended Blocks',FH,fH,center())
    for col in range(2,22): ws.cell(row=13,column=col).fill=FH
    ws.row_dimensions[13].height=11
    for fr,lbl in zip([14,15,16],['F 08:00–10:40','F 10:50–13:30','F 13:40–16:20']):
        apc(ws,fr,1,lbl,FN,Font(color='FF1F3864',bold=True,size=10,name='Arial'),center())
        for rm,sc in RC:
            st2=RSTYLE[rm]
            for di,day in enumerate(DO):
                if day=='F':
                    k=(rm,'F',lbl)
                    if k in occ: apc(ws,fr,sc+di,occ[k],st2['cf'],st2['fn'],center())
                    else:         apc(ws,fr,sc+di,'FREE',FG,fG,center())
                else: apc(ws,fr,sc+di,'—',FGR,fD,center())
        ws.row_dimensions[fr].height=30
    apc(ws,17,1,'Evening Sessions',FH,fH,center())
    for col in range(2,22): ws.cell(row=17,column=col).fill=FH
    ws.row_dimensions[17].height=11
    for ei,lbl in enumerate(['17:55–20:35','18:00–20:00']):
        er=18+ei
        apc(ws,er,1,lbl,FN,Font(color='FF1F3864',bold=True,size=10,name='Arial'),center())
        for rm,sc in RC:
            st2=RSTYLE[rm]
            for di,day in enumerate(DO):
                k=(rm,day,lbl)
                if k in occ: apc(ws,er,sc+di,occ[k],st2['cf'],st2['fn'],center())
                else:         apc(ws,er,sc+di,'FREE',FG,fG,center())
        ws.row_dimensions[er].height=30
    apc(ws,20,1,'F 15:05–17:45',FN,Font(color='FF1F3864',bold=True,size=10,name='Arial'),center())
    for rm,sc in RC:
        st2=RSTYLE[rm]
        for di,day in enumerate(DO):
            if day=='F':
                k=(rm,'F','F 15:05–17:45')
                if k in occ: apc(ws,20,sc+di,occ[k],st2['cf'],st2['fn'],center())
                else:         apc(ws,20,sc+di,'FREE',FG,fG,center())
            else: apc(ws,20,sc+di,'—',FGR,fD,center())
    ws.row_dimensions[20].height=30
    ws.column_dimensions['A'].width=16
    for col in range(2,22): ws.column_dimensions[get_column_letter(col)].width=12

# ── FREE SLOT SUMMARY ──────────────────────────────────────────────────────────
def build_free_slots(wb):
    ws=wb.create_sheet('Free Slot Summary')
    ws.merge_cells('A1:E1')
    apc(ws,1,1,'Fall 2026 — Free Room Time Slots (NB-6.61, NB-6.67, NB-6.6402, NB-L2.79)',FT,fT,left())
    ws.row_dimensions[1].height=20
    ws.merge_cells('A2:E2')
    apc(ws,2,1,'Standard M/Tu/W/Th time blocks and Friday extended blocks with no scheduled section.',FN,fN,left())
    ws.row_dimensions[2].height=15
    for i,h in enumerate(['Room','Day','Block Start','Block End','Notes'],1):
        c=ws.cell(row=3,column=i,value=h); c.fill=FH; c.font=fH; c.alignment=center()
    ws.row_dimensions[3].height=18
    for col,w in zip('ABCDE',[14,6,13,13,18]): ws.column_dimensions[col].width=w

    tr=active[active['Room'].isin(TRACKED_ROOMS)].copy()
    tr['ST']=tr['Mtg Start'].apply(t2time); tr['ET']=tr['Mtg End'].apply(t2time)
    tr['DL']=tr['Days'].apply(parse_days)
    occ_std=set(); occ_fri=set()
    for _,r in tr.iterrows():
        rm=r['Room']; st=r['ST']; et=r['ET']
        if not st: continue
        for day in r['DL']:
            if day!='F':
                for lbl,ss2,se in STD_SLOTS:
                    if st==ss2: occ_std.add((rm,day,ss2))
            else:
                for lbl,ss2,se in STD_SLOTS:
                    if st<=ss2 and et>ss2: occ_std.add((rm,'F',ss2))
                for lbl,fs,fe in FRI_EXT:
                    if st==fs: occ_fri.add((rm,lbl))

    MTH=[('08:00','09:15',dtime(8,0),'Standard slot'),('09:25','10:40',dtime(9,25),'Standard slot'),
         ('10:50','12:05',dtime(10,50),'Standard slot'),('12:15','13:30',dtime(12,15),'Standard slot'),
         ('15:05','16:20',dtime(15,5),'Standard slot'),('16:30','17:45',dtime(16,30),'Standard slot'),
         ('17:55','20:35',dtime(17,55),'Evening'),('18:00','20:00',dtime(18,0),'Evening')]
    FXB=[('08:00','10:40','F 08:00–10:40','Friday extended'),('10:50','13:30','F 10:50–13:30','Friday extended'),
         ('13:40','16:20','F 13:40–16:20','Friday extended'),('15:05','17:45','F 15:05–17:45','Friday extended')]

    row=4; tog=True
    for rm in TRACKED_ROOMS:
        slots=[]
        for day in ['M','Tu','W','Th']:
            for bs,be,bt,note in MTH:
                if (rm,day,bt) not in occ_std: slots.append((rm,day,bs,be,note))
        for bs,be,lbl,note in FXB:
            if (rm,lbl) not in occ_fri: slots.append((rm,'F',bs,be,note))
        for sl in slots:
            rf=FA if tog else FW; tog=not tog
            for ci,v in enumerate(sl,1):
                c=ws.cell(row=row,column=ci,value=v); c.fill=rf; c.font=fB; c.alignment=center()
            ws.row_dimensions[row].height=15; row+=1

# ── BUILD WORKBOOK ─────────────────────────────────────────────────────────────
# Load existing workbook to preserve Config; rebuild analysis sheets
if os.path.exists(WORKBOOK_PATH):
    wb = load_workbook(WORKBOOK_PATH)
    # Remove all sheets except Config
    for sname in list(wb.sheetnames):
        if sname != 'Config':
            del wb[sname]
else:
    wb = Workbook()
    wb.remove(wb.active)

# Build Config first (create if missing)
if 'Config' not in wb.sheetnames:
    build_config(wb)

# Build analysis sheets
build_capacity(wb)
build_section_detail(wb)
build_grid(wb)
build_free_slots(wb)

# Sheet order: Config first, then analysis
sheet_order = ['Config','Capacity by Course','Section Detail','Room Availability Grid','Free Slot Summary']
wb._sheets = [wb[s] for s in sheet_order]

wb.save(WORKBOOK_PATH)
print(f'Saved  : {WORKBOOK_PATH}')
print(f'Active : {len(active)} sections  |  Stop Enrl: {len(stop_enrl)}')
